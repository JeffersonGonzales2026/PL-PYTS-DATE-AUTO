import calendar
import re
from datetime import date, datetime, timedelta
from io import BytesIO
from numbers import Number
from pathlib import Path
from zipfile import BadZipFile

import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from openpyxl.utils.datetime import from_excel
from openpyxl.utils.exceptions import InvalidFileException

try:
    import msoffcrypto
except ImportError:
    msoffcrypto = None


st.set_page_config(
    page_title="PL Payment & Date Automation",
    page_icon="PL",
    layout="wide",
)


OOXML_FILE_SIGNATURE = b"PK\x03\x04"
OLE_FILE_SIGNATURE = b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"

DEFAULT_SHEET_PREFIX = "payment as of"
DEFAULT_LAN_COL = "C"
DEFAULT_AMOUNT_COL = "E"
DEFAULT_DATE_COL = "F"
DATA_START_ROW = 2
OUTPUT_SHEET_NAME = "Sheet1"


def normalize_lan(value):
    if value is None:
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def parse_excel_date(value):
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, (int, float)):
        try:
            converted = from_excel(value)
            if isinstance(converted, datetime):
                return converted.date()
            if isinstance(converted, date):
                return converted
        except Exception:
            return None

    if isinstance(value, str):
        text = value.strip()
        if not text:
            return None

        for fmt in (
            "%m/%d/%Y",
            "%m-%d-%Y",
            "%d/%m/%Y",
            "%d-%m-%Y",
            "%Y-%m-%d",
            "%b %d, %Y",
            "%B %d, %Y",
        ):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                pass

    return None


def month_str_to_num(value):
    month_map = {
        "jan": 1,
        "january": 1,
        "feb": 2,
        "february": 2,
        "mar": 3,
        "march": 3,
        "apr": 4,
        "april": 4,
        "may": 5,
        "jun": 6,
        "june": 6,
        "jul": 7,
        "july": 7,
        "aug": 8,
        "august": 8,
        "sep": 9,
        "sept": 9,
        "september": 9,
        "oct": 10,
        "october": 10,
        "nov": 11,
        "november": 11,
        "dec": 12,
        "december": 12,
    }
    return month_map.get(str(value).strip().lower())


def find_sheet_by_prefix(workbook, sheet_prefix):
    prefix = sheet_prefix.strip().lower()
    for name in workbook.sheetnames:
        if name.strip().lower().startswith(prefix):
            return name
    return None


def extract_month_year_from_sheetname(sheet_name, sheet_prefix=None):
    text = str(sheet_name).strip()
    if sheet_prefix:
        text = re.sub(
            rf"(?i)^\s*{re.escape(sheet_prefix)}\s*",
            "",
            text,
        ).strip()

    parts = re.split(r"[\s_-]+", text)
    for index in range(len(parts) - 1):
        month = month_str_to_num(parts[index])
        try:
            year = int(parts[index + 1])
        except ValueError:
            year = None
        if month and year:
            return month, year

    today = datetime.today()
    return today.month, today.year


def first_monday(year, month):
    current = date(year, month, 1)
    while current.weekday() != 0:
        current += timedelta(days=1)
    return current


def build_week_ranges(year, month):
    ranges = []
    last_day = calendar.monthrange(year, month)[1]
    for start_day in range(1, last_day + 1, 7):
        start = date(year, month, start_day)
        end_day = min(start_day + 6, last_day)
        end = date(year, month, end_day)
        ranges.append((start, end))
    return ranges


def week_bucket(payment_date, year, month, total_weeks):
    if payment_date.year != year or payment_date.month != month:
        return None
    week_number = ((payment_date.day - 1) // 7) + 1
    return min(week_number, total_weeks)


def add_amount_to_cell(worksheet, cell_address, amount):
    current = worksheet[cell_address].value
    worksheet[cell_address] = amount if current in (None, "") else current + amount


def set_date_cell(worksheet, cell_address, payment_date):
    worksheet[cell_address] = payment_date
    worksheet[cell_address].number_format = "mm/dd/yyyy"


def keep_latest_date(worksheet, cell_address, new_date):
    current_date = parse_excel_date(worksheet[cell_address].value)
    if current_date is None or new_date > current_date:
        set_date_cell(worksheet, cell_address, new_date)


def label_week(index, week_ranges):
    start, end = week_ranges[index]
    return f"Payment week {index + 1} ({start.strftime('%b %d')} to {end.strftime('%b %d')})"


def normalize_header(value):
    if value is None:
        return ""
    return str(value).strip().lower()


def find_header_column(headers, required_parts):
    for name, col in headers.items():
        if all(part in name for part in required_parts):
            return col
    return None


def get_header_map(worksheet):
    return {
        normalize_header(worksheet.cell(row=1, column=col).value): get_column_letter(col)
        for col in range(1, worksheet.max_column + 1)
        if normalize_header(worksheet.cell(row=1, column=col).value)
    }


def current_month_year_from_sheet(sheet_name=None):
    if sheet_name:
        month, year = extract_month_year_from_sheetname(sheet_name)
        if month and year:
            return month, year
    today = datetime.today()
    return today.month, today.year


def payment_amount_header(index, week_ranges):
    start, end = week_ranges[index]
    return f"Payment Amount (Week {index + 1}: {start.strftime('%b %d')} - {end.strftime('%b %d')})"


def payment_date_header(index):
    return f"Payment Date (Week {index + 1})"


def parse_week_number_from_header(header):
    header = normalize_header(header)
    if not header:
        return None
    match = re.search(r"week\s*([1-5])", header)
    if match:
        return int(match.group(1))
    match = re.search(r"week[_\s]?([1-5])", header)
    if match:
        return int(match.group(1))
    return None


def parse_target_month_year(worksheet):
    month_keywords = r"\b(january|february|march|april|may|june|july|august|september|october|november|december|jan|feb|mar|apr|jun|jul|aug|sep|sept|oct|nov|dec)\b"
    title_month, title_year = extract_month_year_from_sheetname(worksheet.title)
    if title_month and title_year and re.search(month_keywords, worksheet.title, flags=re.IGNORECASE):
        return title_month, title_year

    for cell in worksheet[1]:
        value = normalize_header(cell.value)
        if not value:
            continue

        month_match = re.search(month_keywords, value)
        if month_match:
            month_num = month_str_to_num(month_match.group(1))
            year_match = re.search(r"(20\d{2}|19\d{2})", value)
            if month_num:
                return month_num, int(year_match.group(1)) if year_match else datetime.today().year

    return datetime.today().month, datetime.today().year


def detect_week_columns(worksheet):
    headers = get_header_map(worksheet)
    amount_map = {}
    date_map = {}

    for header, col in headers.items():
        week_number = parse_week_number_from_header(header)
        if not week_number:
            continue

        if "payment" in header and "amount" in header:
            amount_map[week_number] = col
        elif "payment" in header and "date" in header:
            date_map[week_number] = col

    week_numbers = sorted(set(amount_map.keys()) | set(date_map.keys()))
    if not week_numbers:
        raise ValueError(
            "Could not find existing weekly payment/date columns in the tracker workbook. "
            "Please upload the workbook that already contains the week payment and date columns."
        )

    missing = []
    week_columns = []
    for week_number in week_numbers:
        amount_col = amount_map.get(week_number)
        date_col = date_map.get(week_number)
        if not amount_col:
            missing.append(f"Payment Amount Week {week_number}")
        if not date_col:
            missing.append(f"Payment Date Week {week_number}")
        if amount_col and date_col:
            week_columns.append((amount_col, date_col))

    if missing:
        raise ValueError(
            "Target workbook is missing these required columns: " + ", ".join(missing)
        )

    return week_columns


def parse_amount(value):
    if value is None or value == "":
        return None
    if isinstance(value, Number):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def update_target_sheet_with_source_data(source, target, source_account_col, source_amount_col, source_date_col, target_account_col):
    target_account_map = {}
    for row in range(DATA_START_ROW, target.max_row + 1):
        account_key = normalize_lan(target[f"{target_account_col}{row}"].value)
        if account_key:
            target_account_map[account_key] = row

    month, year = parse_target_month_year(target)
    week_ranges = build_week_ranges(year, month)
    week_columns = detect_week_columns(target)

    skipped_rows = 0
    missing_accounts = 0

    for row in range(DATA_START_ROW, source.max_row + 1):
        account_key = normalize_lan(source[f"{source_account_col}{row}"].value)
        amount = parse_amount(source[f"{source_amount_col}{row}"].value)
        payment_date = parse_excel_date(source[f"{source_date_col}{row}"].value)

        if not account_key or amount is None or amount == 0 or payment_date is None:
            skipped_rows += 1
            continue

        week_number = week_bucket(payment_date, year, month, len(week_ranges))
        if week_number is None:
            skipped_rows += 1
            continue

        target_row = target_account_map.get(account_key)
        if target_row is None:
            missing_accounts += 1
            continue

        amount_col, date_col = week_columns[week_number - 1]
        add_amount_to_cell(target, f"{amount_col}{target_row}", amount)
        keep_latest_date(target, f"{date_col}{target_row}", payment_date)

    total_col = find_header_column(get_header_map(target), ["total", "payment"])
    if total_col:
        total_index = column_index_from_string(total_col)
        total_columns = [col for col, _ in week_columns]
        for row in range(DATA_START_ROW, target.max_row + 1):
            formula_parts = ",".join(f"{column}{row}" for column in total_columns)
            target.cell(row=row, column=total_index, value=f"=SUM({formula_parts})")
        style_output_sheet(target, total_index, target.max_row)
    else:
        style_output_sheet(target, target.max_column, target.max_row)
    return {
        "month": calendar.month_name[month],
        "year": year,
        "processed_rows": source.max_row - DATA_START_ROW + 1,
        "skipped_rows": skipped_rows,
        "missing_accounts": missing_accounts,
        "week_count": len(week_ranges),
    }


def style_output_sheet(worksheet, total_col, last_row):
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    total_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    white_bold = Font(color="FFFFFF", bold=True)
    bold = Font(bold=True)
    centered = Alignment(horizontal="center", vertical="center", wrap_text=True)

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for col in range(1, total_col + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = white_bold if col != total_col else bold
        cell.fill = total_fill if col == total_col else header_fill
        cell.alignment = centered

    for row in worksheet.iter_rows(min_row=1, max_row=max(last_row, 1), min_col=1, max_col=total_col):
        for cell in row:
            cell.alignment = centered

    for col in range(1, total_col + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for cell in worksheet[column_letter]:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 35)


def rebuild_output_sheet(workbook, sheet_prefix, lan_col, amount_col, date_col):
    source_sheet_name = find_sheet_by_prefix(workbook, sheet_prefix)
    if not source_sheet_name:
        raise ValueError(f"No sheet found that starts with '{sheet_prefix}'.")

    source = workbook[source_sheet_name]
    month, year = extract_month_year_from_sheetname(source_sheet_name, sheet_prefix)
    week_ranges = build_week_ranges(year, month)
    week_count = len(week_ranges)

    if OUTPUT_SHEET_NAME in workbook.sheetnames:
        output_index = workbook.sheetnames.index(OUTPUT_SHEET_NAME)
        workbook.remove(workbook[OUTPUT_SHEET_NAME])
        output = workbook.create_sheet(OUTPUT_SHEET_NAME, output_index)
    else:
        output = workbook.create_sheet(OUTPUT_SHEET_NAME)

    output["A1"] = "LAN2"
    for index in range(week_count):
        amount_output_col = 2 + (index * 2)
        date_output_col = amount_output_col + 1
        output.cell(row=1, column=amount_output_col, value=label_week(index, week_ranges))
        output.cell(row=1, column=date_output_col, value=f"Payment Date (Week {index + 1})")

    total_col = 2 + (week_count * 2)
    output.cell(row=1, column=total_col, value="TOTAL")

    lan_to_output_row = {}
    next_output_row = 2
    skipped_rows = 0

    for row in range(DATA_START_ROW, source.max_row + 1):
        lan_key = normalize_lan(source[f"{lan_col}{row}"].value)
        amount = source[f"{amount_col}{row}"].value
        payment_date = parse_excel_date(source[f"{date_col}{row}"].value)

        if not lan_key or payment_date is None or amount in (None, "", 0):
            skipped_rows += 1
            continue

        if lan_key not in lan_to_output_row:
            lan_to_output_row[lan_key] = next_output_row
            output[f"A{next_output_row}"] = lan_key
            next_output_row += 1

        target_row = lan_to_output_row[lan_key]
        week_number = week_bucket(payment_date, month, year, week_count)
        if week_number is None:
            skipped_rows += 1
            continue

        amount_column = get_column_letter(2 + ((week_number - 1) * 2))
        date_column = get_column_letter(3 + ((week_number - 1) * 2))

        add_amount_to_cell(output, f"{amount_column}{target_row}", amount)
        keep_latest_date(output, f"{date_column}{target_row}", payment_date)

    amount_columns = [get_column_letter(2 + (index * 2)) for index in range(week_count)]
    for row in range(2, next_output_row):
        formula_parts = ",".join(f"{column}{row}" for column in amount_columns)
        output.cell(row=row, column=total_col, value=f"=SUM({formula_parts})")

    style_output_sheet(output, total_col, next_output_row - 1)

    return {
        "source_sheet": source_sheet_name,
        "month": calendar.month_name[month],
        "year": year,
        "account_count": len(lan_to_output_row),
        "skipped_rows": skipped_rows,
        "week_count": week_count,
    }


def inspect_uploaded_excel(file_name, raw_bytes):
    normalized_name = Path(file_name or "uploaded workbook").name
    lowered_name = normalized_name.lower()
    header_bytes = raw_bytes[:2048]
    stripped_header = header_bytes.lstrip().lower()

    if lowered_name.startswith("~$"):
        return "excel_temp_lock"
    if not raw_bytes:
        return "empty"
    if raw_bytes.startswith(OOXML_FILE_SIGNATURE):
        return "ooxml_zip"
    if raw_bytes.startswith(OLE_FILE_SIGNATURE):
        return "ole_compound"
    if stripped_header.startswith((b"<!doctype html", b"<html")):
        return "html"
    if b"\x00" not in header_bytes:
        has_delimiter = any(delimiter in header_bytes for delimiter in (b",", b";", b"\t"))
        has_line_break = any(line_break in header_bytes for line_break in (b"\r", b"\n"))
        if has_delimiter and has_line_break:
            return "delimited_text"
    return "unknown"


def build_upload_precheck_error(file_name, file_type):
    if file_type == "excel_temp_lock":
        return (
            f'"{file_name}" looks like Excel\'s temporary lock file (~$...), not the actual workbook. '
            "Close the source workbook if it is open, then upload the real file."
        )
    if file_type == "empty":
        return f'"{file_name}" is empty. Please upload a valid .xlsx or .xlsm workbook.'
    if file_type == "html":
        return f'"{file_name}" is an HTML page, not an Excel workbook.'
    if file_type == "delimited_text":
        return f'"{file_name}" looks like a CSV/text file. Please upload a valid Excel workbook.'
    return None


def format_workbook_open_error(error):
    if isinstance(error, BadZipFile):
        return "File is not a valid .xlsx/.xlsm workbook archive."
    if isinstance(error, InvalidFileException):
        return str(error).strip() or "Unsupported workbook format."
    return str(error).strip() or error.__class__.__name__


def open_uploaded_workbook(uploaded_file, password=None):
    uploaded_file.seek(0)
    raw = uploaded_file.read()
    file_name = Path(getattr(uploaded_file, "name", "uploaded workbook")).name
    file_type = inspect_uploaded_excel(file_name, raw)
    precheck_error = build_upload_precheck_error(file_name, file_type)

    if precheck_error:
        raise ValueError(precheck_error)

    try:
        return load_workbook(BytesIO(raw))
    except Exception as direct_error:
        if not password:
            raise ValueError(
                "The workbook could not be opened normally. If it is password-protected, enter the password and try again. "
                f"Details: {format_workbook_open_error(direct_error)}"
            )

        if msoffcrypto is None:
            raise ValueError("Password-protected workbooks require the msoffcrypto-tool package.")

        try:
            decrypted = BytesIO()
            office_file = msoffcrypto.OfficeFile(BytesIO(raw))
            office_file.load_key(password=password)
            office_file.decrypt(decrypted)
            decrypted.seek(0)
            return load_workbook(decrypted)
        except Exception as password_error:
            raise ValueError(
                "The workbook still could not be opened using the provided password. "
                f"Details: {format_workbook_open_error(password_error)}"
            ) from password_error


def run():
    st.title("PL Payment & Date Automation")
    st.caption(
        "Upload the updated payment source file and your tracker file to auto-fill Week 1 to Week 5 payment amounts and dates."
    )

    with st.sidebar:
        st.header("Source workbook settings")
        source_sheet_prefix = st.text_input("Source sheet starts with", "")
        source_account_col = st.text_input("Source account column", DEFAULT_LAN_COL).strip().upper()
        source_amount_col = st.text_input("Source amount column", DEFAULT_AMOUNT_COL).strip().upper()
        source_date_col = st.text_input("Source payment date column", DEFAULT_DATE_COL).strip().upper()

        st.header("Target tracker settings")
        target_sheet_prefix = st.text_input("Target sheet name or prefix", "")
        target_account_col = st.text_input("Target account column", DEFAULT_LAN_COL).strip().upper()

        st.header("Protected workbook passwords")
        source_password = st.text_input("Source workbook password", type="password")
        target_password = st.text_input("Tracker workbook password", type="password")

    source_file = st.file_uploader("Upload updated payments workbook", type=["xlsx", "xlsm"], key="source")
    target_file = st.file_uploader("Upload tracker workbook to fill", type=["xlsx", "xlsm"], key="target")

    if source_file is None or target_file is None:
        st.info("Upload both the updated payments workbook and the tracker workbook to update week amount/date fields.")
        return

    try:
        source_workbook = open_uploaded_workbook(source_file, password=source_password or None)
        target_workbook = open_uploaded_workbook(target_file, password=target_password or None)

        source_sheet_name = source_workbook.sheetnames[0]
        if source_sheet_prefix:
            matched = find_sheet_by_prefix(source_workbook, source_sheet_prefix)
            if matched:
                source_sheet_name = matched
            else:
                raise ValueError(f"No source sheet found that starts with '{source_sheet_prefix}'.")

        target_sheet_name = target_workbook.sheetnames[0]
        if target_sheet_prefix:
            matched = find_sheet_by_prefix(target_workbook, target_sheet_prefix)
            if matched:
                target_sheet_name = matched
            else:
                raise ValueError(f"No target sheet found that starts with '{target_sheet_prefix}'.")

        source_sheet = source_workbook[source_sheet_name]
        target_sheet = target_workbook[target_sheet_name]

        summary = update_target_sheet_with_source_data(
            source_sheet,
            target_sheet,
            source_account_col,
            source_amount_col,
            source_date_col,
            target_account_col,
        )

        output = BytesIO()
        target_workbook.save(output)
        output.seek(0)

        col1, col2, col3, col4 = st.columns(4)
        col1.metric("Source sheet", source_sheet_name)
        col2.metric("Target sheet", target_sheet_name)
        col3.metric("Processed rows", summary["processed_rows"])
        col4.metric("Weeks", summary["week_count"])

        if summary["skipped_rows"]:
            st.caption(
                f'{summary["skipped_rows"]} source row(s) were skipped because the account, amount, or payment date was blank/invalid, or outside the target month.'
            )
        if summary["missing_accounts"]:
            st.caption(
                f'{summary["missing_accounts"]} payment row(s) could not be matched to account numbers in the target workbook.'
            )

        st.success(
            "Done. Your tracker workbook has been updated with weekly payment amounts, dates, and totals."
        )
        st.download_button(
            "Download Updated Tracker Workbook",
            data=output,
            file_name="PL_PAYMENT_TRACKER_UPDATED.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except Exception as error:
        st.error(f"Error: {error}")


if __name__ == "__main__":
    run()
